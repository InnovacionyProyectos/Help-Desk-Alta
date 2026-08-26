"""Excel de tickets — equivalente al export con ExcelJS de
reports.service.ts. 13 columnas, encabezado en negrita, auto-filtro
A1:M1."""

import io

from openpyxl import Workbook
from openpyxl.styles import Font

from app.models.ticket import Ticket
from app.services.reports_service import PRIORITY_LABELS, STATUS_LABELS

# Tipografía de marca: a diferencia del PDF (reportlab embebe el .ttf, ver
# app/exports/fonts.py), un .xlsx solo declara el NOMBRE de la fuente —
# Excel la sustituye por una fuente local si "DM Sans" no está instalada
# en el equipo que lo abre. Aun así es más correcto declararla que dejar
# el Calibri por defecto de openpyxl, que no tiene relación con la marca.
BRAND_FONT = "DM Sans"

COLUMNS = [
    "Ticket#",
    "Asunto",
    "Categoría",
    "Subcategoría",
    "Tipificación",
    "Prioridad",
    "Estado",
    "Solicitante",
    "Asignado a",
    "Área",
    "Creado",
    "Resuelto",
    "Cerrado",
]


def _fmt_date(dt) -> str:
    return dt.strftime("%Y-%m-%d %H:%M") if dt else ""


def build_tickets_excel(tickets: list[Ticket]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Tickets"

    ws.append(COLUMNS)
    for cell in ws[1]:
        cell.font = Font(name=BRAND_FONT, bold=True)
    ws.auto_filter.ref = "A1:M1"

    for t in tickets:
        ws.append(
            [
                t.ticket_number,
                t.subject,
                t.category.name if t.category else "",
                t.subcategory.name if t.subcategory else "",
                t.typification.name if t.typification else "",
                PRIORITY_LABELS.get(t.priority, t.priority),
                STATUS_LABELS.get(t.status.code, t.status.name),
                t.requester.full_name,
                t.assigned_to.full_name if t.assigned_to else "",
                t.assigned_area.name if t.assigned_area else "",
                _fmt_date(t.created_at),
                _fmt_date(t.resolved_at),
                _fmt_date(t.closed_at),
            ]
        )
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name=BRAND_FONT)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
