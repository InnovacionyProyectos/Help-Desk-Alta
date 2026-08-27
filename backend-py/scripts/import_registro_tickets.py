"""Script de una sola vez para importar el histórico real de tickets desde
`Plantilla Mesa de Servicios TI 2026.xlsx` (hoja `Registro_Tickets`, 535
filas) a la base de datos de producción (`helpdesk`). No forma parte de la
app, no se monta en ninguna ruta — se corre a mano desde la terminal.

Ver el plan completo (mapeo de campos, decisiones tomadas con el usuario)
en `C:\\Users\\proyectos\\.claude\\plans\\refactored-percolating-unicorn.md`,
sección "Carga histórica: Registro_Tickets -> producción real".

Uso:
  python scripts/import_registro_tickets.py            # dry-run (default, no toca la BD)
  python scripts/import_registro_tickets.py --execute   # escribe de verdad, todo en una transacción
"""

import argparse
import asyncio
import difflib
import sys
import unicodedata
import uuid
from datetime import datetime, timedelta, timezone
from pathlib import Path

if sys.platform == "win32":
    asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())

import openpyxl
from sqlalchemy import select, text
from sqlalchemy.orm import selectinload

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from app.database import SessionLocal  # noqa: E402
from app.models.area import Area  # noqa: E402
from app.models.audit import AuditLog  # noqa: E402
from app.models.classification import TicketSubcategory, TicketTypification  # noqa: E402
from app.models.role import Role  # noqa: E402
from app.models.ticket import (  # noqa: E402
    Ticket,
    TicketAssignmentHistory,
    TicketComment,
    TicketStatus,
    TicketStatusHistory,
)
from app.models.user import User  # noqa: E402
from app.security.passwords import hash_password  # noqa: E402

EXCEL_PATH = Path(__file__).resolve().parent.parent.parent / "Plantilla Mesa de Servicios TI 2026.xlsx"
SHEET = "Registro_Tickets"

AREA_MAP = {
    "administrativa financiera": "Administrativa y Financiera",
    "credito": "Crédito",
    "innovacion y proyectos": "Innovación y Proyectos",
    "alianzas estrategicas": "Alianzas Estratégicas",
    "operaciones": "Operaciones",
    "comercial": "Comercial",
    "gerencia general": "Gerencia General",
    "mercadeo": "Mercadeo",
}

ANALISTA_EMAIL = {
    "andres rubiano": "mrubiano@altatefinancia.com",
    "carolina marin": "cmarin@altatefinancia.com",
}

# Nombres sin cuenta -> decisión tomada con el usuario (ver plan).
NEW_USERS = {
    "eider javier garzon": {
        "email": "egarzon@altatefinancia.com",
        "first_name": "Eider Javier",
        "last_name": "Garzon",
    },
    "yenifer suarez": {
        "email": "ysuarez@altatefinancia.com",
        "first_name": "Yenifer",
        "last_name": "Suarez",
    },
    "lorena rueda": {
        "email": "lrueda@altatefinancia.com",
        "first_name": "Lorena",
        "last_name": "Rueda Agredo",
    },
}
REASSIGN_TO_CGIL = {"yamileth - contadora externa", "yohanna gonzalez - contadora"}
CGIL_EMAIL = "cgil@altatefinancia.com"

# Variantes de ortografía del mismo nombre que el matching por tokens no
# resuelve solo (ej. "Leidy Johanna" en el Excel vs "Laidy Johana" en BD) —
# confirmado que es la misma persona antes de agregar la excepción.
NAME_OVERRIDE = {
    "leidy johanna gomez": "jgomez@altatefinancia.com",
}

# Tipificación pedida en el Excel que no existe en el catálogo (1 ticket,
# fila 55) — decisión tomada con el usuario: usar la más cercana ya
# existente en la misma subcategoría en vez de crear una tipificación nueva
# para un solo caso histórico.
CHAIN_OVERRIDE = {
    ("equipos", "sistema_operativo", "instalacion de sistema operativo"): "instalacion de aplicaciones",
}

STATUS_MAP = {"cerrado": "CLOSED", "en proceso (soporte l2)": "IN_PROGRESS"}
TIPO_MAP = {"requerimiento": "REQUERIMIENTO", "incidente": "INCIDENTE"}
IMPACTO_MAP = {"alta": "HIGH", "media": "MEDIUM", "baja": "LOW"}

MIGRATION_REASON = "Migración de histórico Mesa de Servicios TI 2026"
FUZZY_THRESHOLD = 0.72


def norm(s) -> str:
    if s is None:
        return ""
    s = str(s).strip().lower().rstrip(".")
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode()
    return " ".join(s.split())


def read_rows() -> list[dict]:
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb[SHEET]
    rows = []
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value in (None, ""):
            continue
        rows.append(
            {
                "excel_row": r,
                "created_at": ws.cell(row=r, column=3).value,
                "canal": ws.cell(row=r, column=4).value,
                "usuario": ws.cell(row=r, column=5).value,
                "empresa": ws.cell(row=r, column=6).value,
                "area": ws.cell(row=r, column=7).value,
                "tipo": ws.cell(row=r, column=8).value,
                "categoria": ws.cell(row=r, column=9).value,
                "subcategoria": ws.cell(row=r, column=10).value,
                "tipificacion": ws.cell(row=r, column=11).value,
                "impacto": ws.cell(row=r, column=12).value,
                "inicio_atencion": ws.cell(row=r, column=13).value,
                "cierre": ws.cell(row=r, column=14).value,
                "estado": ws.cell(row=r, column=15).value,
                "observaciones": ws.cell(row=r, column=16).value,
                "analista": ws.cell(row=r, column=17).value,
                "solucion": ws.cell(row=r, column=18).value,
            }
        )
    return rows


COLOMBIA_TZ = timezone(timedelta(hours=-5))


def as_utc(dt) -> datetime:
    """Las fechas del Excel llegan naive (sin tz) — se asumen hora local de
    Colombia (UTC-5) y se guardan en UTC, igual que hace la app en vivo con
    `datetime.now(timezone.utc)`."""
    if dt.tzinfo is not None:
        return dt.astimezone(timezone.utc)
    return dt.replace(tzinfo=COLOMBIA_TZ).astimezone(timezone.utc)


async def build_lookups(db):
    result = await db.execute(select(Area))
    areas_by_name = {a.name: a for a in result.scalars().all()}

    result = await db.execute(select(User).where(User.deleted_at.is_(None)))
    users = list(result.scalars().all())
    users_by_email = {u.email.lower(): u for u in users}
    users_tokenized = [
        {"user": u, "tokens": set(norm(f"{u.first_name} {u.last_name}").split())} for u in users
    ]

    result = await db.execute(
        select(TicketTypification)
        .where(TicketTypification.is_active.is_(True))
        .options(selectinload(TicketTypification.subcategory).selectinload(TicketSubcategory.category))
    )
    typifications = list(result.scalars().unique().all())
    chain_exact: dict[tuple[str, str, str], TicketTypification] = {}
    by_cat_sub: dict[tuple[str, str], list[tuple[str, TicketTypification]]] = {}
    for t in typifications:
        if not (t.subcategory.is_active and t.subcategory.category.is_active):
            continue
        nc = norm(t.subcategory.category.name)
        ns = norm(t.subcategory.name)
        nt = norm(t.name)
        chain_exact[(nc, ns, nt)] = t
        by_cat_sub.setdefault((nc, ns), []).append((nt, t))

    result = await db.execute(select(TicketStatus))
    statuses_by_code = {s.code: s for s in result.scalars().all()}

    return {
        "areas_by_name": areas_by_name,
        "users_by_email": users_by_email,
        "users_tokenized": users_tokenized,
        "chain_exact": chain_exact,
        "by_cat_sub": by_cat_sub,
        "statuses_by_code": statuses_by_code,
    }


def resolve_typification(row, lookups):
    nc = norm(row["categoria"])
    ns = norm(row["subcategoria"])
    nt = norm(row["tipificacion"])
    nt = CHAIN_OVERRIDE.get((nc, ns, nt), nt)

    exact = lookups["chain_exact"].get((nc, ns, nt))
    if exact:
        return exact, "exact"

    candidates = lookups["by_cat_sub"].get((nc, ns), [])
    for cand_norm, cand_obj in candidates:
        if nt in cand_norm or cand_norm in nt:
            return cand_obj, "substring"

    best_obj, best_ratio = None, 0.0
    for cand_norm, cand_obj in candidates:
        ratio = difflib.SequenceMatcher(None, nt, cand_norm).ratio()
        if ratio > best_ratio:
            best_ratio, best_obj = ratio, cand_obj
    if best_obj is not None and best_ratio >= FUZZY_THRESHOLD:
        return best_obj, f"fuzzy:{best_ratio:.2f}"

    return None, "unresolved"


def resolve_requester(row, lookups):
    key = norm(row["usuario"])
    if key in NEW_USERS:
        return "new", NEW_USERS[key]["email"]
    if key in REASSIGN_TO_CGIL:
        return "cgil", CGIL_EMAIL
    if key in NAME_OVERRIDE:
        return "override", NAME_OVERRIDE[key]

    tokens = set(key.split())
    matches = [t["user"] for t in lookups["users_tokenized"] if tokens.issubset(t["tokens"])]
    if len(matches) == 1:
        return "match", matches[0].email
    return "unresolved", matches


async def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--execute", action="store_true", help="Escribe de verdad (default: dry-run)")
    args = parser.parse_args()

    rows = read_rows()
    print(f"Filas leídas del Excel: {len(rows)}")

    async with SessionLocal() as db:
        lookups = await build_lookups(db)

        resolution_counts = {"exact": 0, "substring": 0, "unresolved_chain": 0}
        fuzzy_examples = []
        unresolved_chains = []
        unresolved_requesters = []
        requester_email_by_row = {}

        for row in rows:
            typ, method = resolve_typification(row, lookups)
            if typ is None:
                resolution_counts["unresolved_chain"] += 1
                unresolved_chains.append(row)
            elif method == "exact":
                resolution_counts["exact"] += 1
            elif method == "substring":
                resolution_counts["substring"] += 1
            else:
                fuzzy_examples.append((row["excel_row"], row["categoria"], row["subcategoria"], row["tipificacion"], typ.name, method))

            kind, ident = resolve_requester(row, lookups)
            if kind == "unresolved":
                unresolved_requesters.append((row["excel_row"], row["usuario"], ident))
            else:
                requester_email_by_row[row["excel_row"]] = ident

        print("\n--- Clasificación (categoría/subcategoría/tipificación) ---")
        print(f"  exacto: {resolution_counts['exact']}")
        print(f"  substring: {resolution_counts['substring']}")
        print(f"  fuzzy: {len(fuzzy_examples)}")
        print(f"  SIN RESOLVER: {resolution_counts['unresolved_chain']}")
        if fuzzy_examples:
            print("\n  Ejemplos de match difuso (revisar):")
            for ex in fuzzy_examples:
                print("   fila", ex[0], "-", ex[1], "/", ex[2], "/", repr(ex[3]), "->", repr(ex[4]), f"({ex[5]})")
        if unresolved_chains:
            print("\n  Filas SIN clasificación resuelta (no se puede continuar sin decidir esto):")
            for r in unresolved_chains:
                print("   fila", r["excel_row"], "-", r["categoria"], "/", r["subcategoria"], "/", repr(r["tipificacion"]))

        print("\n--- Solicitantes ---")
        print(f"  resueltos: {len(requester_email_by_row)}")
        print(f"  SIN RESOLVER: {len(unresolved_requesters)}")
        if unresolved_requesters:
            print("\n  Nombres sin cuenta y sin regla en el script (no cubiertos por el plan):")
            for excel_row, name, matches in unresolved_requesters:
                print("   fila", excel_row, "-", repr(name), "candidatos ambiguos:", len(matches))

        blocking = bool(unresolved_chains) or bool(unresolved_requesters)
        if blocking:
            print("\n*** HAY FILAS SIN RESOLVER. No se puede continuar (ni en dry-run del todo, ni en --execute). ***")
            return

        print("\nTodo resuelto correctamente. 0 filas bloqueantes.")

        if not args.execute:
            print("\n(dry-run — no se escribió nada. Corre con --execute para importar de verdad.)")
            return

        print("\n=== EJECUTANDO IMPORTACIÓN REAL ===")

        # --- 1. Crear los 3 usuarios nuevos (idempotente: si ya existen, se reusan) ---
        end_user_role = (await db.execute(select(Role).where(Role.code == "END_USER"))).scalar_one()
        for key, info in NEW_USERS.items():
            existing = lookups["users_by_email"].get(info["email"].lower())
            if existing:
                continue
            new_user = User(
                email=info["email"],
                password_hash=hash_password(uuid.uuid4().hex + uuid.uuid4().hex),
                first_name=info["first_name"],
                last_name=info["last_name"],
                role_id=end_user_role.id,
                area_id=None,
            )
            db.add(new_user)
            await db.flush()
            db.add(
                AuditLog(
                    user_id=None,
                    action="CREATE",
                    entity="User",
                    entity_id=str(new_user.id),
                    new_values={"email": new_user.email, "reason": "Creado por import_registro_tickets.py"},
                )
            )
            lookups["users_by_email"][info["email"].lower()] = new_user
            lookups["users_tokenized"].append(
                {"user": new_user, "tokens": set(norm(f"{new_user.first_name} {new_user.last_name}").split())}
            )
            print("  usuario creado:", info["email"])

        await db.flush()

        # --- 2. Insertar los 535 tickets en orden cronológico ---
        rows_sorted = sorted(rows, key=lambda r: r["created_at"])
        open_status = lookups["statuses_by_code"]["OPEN"]

        for row in rows_sorted:
            typification, _ = resolve_typification(row, lookups)
            subcategory = typification.subcategory
            category = subcategory.category

            requester_email = requester_email_by_row[row["excel_row"]]
            requester = lookups["users_by_email"][requester_email.lower()]

            analista_email = ANALISTA_EMAIL[norm(row["analista"])]
            analista = lookups["users_by_email"][analista_email.lower()]

            area_name = AREA_MAP[norm(row["area"])]
            area = lookups["areas_by_name"][area_name]

            status_code = STATUS_MAP[norm(row["estado"])]
            final_status = lookups["statuses_by_code"][status_code]

            created_at = as_utc(row["created_at"])
            inicio_atencion = as_utc(row["inicio_atencion"])
            cierre = as_utc(row["cierre"]) if row["cierre"] else None

            seq = (await db.execute(text("SELECT nextval('ticket_number_seq')"))).scalar_one()
            ticket_number = f"HD-{datetime.now().year}-{str(seq).zfill(5)}"

            ticket = Ticket(
                id=uuid.uuid4(),
                ticket_number=ticket_number,
                subject=row["tipificacion"].strip(),
                description=str(row["observaciones"]).strip(),
                category_id=category.id,
                subcategory_id=subcategory.id,
                typification_id=typification.id,
                status_id=final_status.id,
                priority=IMPACTO_MAP[norm(row["impacto"])],
                ticket_type=TIPO_MAP[norm(row["tipo"])],
                requester_id=requester.id,
                assigned_to_id=analista.id,
                assigned_area_id=area.id,
                resolved_at=cierre if status_code == "CLOSED" else None,
                closed_at=cierre if status_code == "CLOSED" else None,
                created_at=created_at,
                updated_at=cierre or inicio_atencion,
            )
            db.add(ticket)
            await db.flush()

            db.add(
                TicketAssignmentHistory(
                    ticket_id=ticket.id,
                    from_user_id=None,
                    to_user_id=analista.id,
                    assigned_by=analista.id,
                    reason=None,
                    created_at=inicio_atencion,
                )
            )
            db.add(
                TicketStatusHistory(
                    ticket_id=ticket.id,
                    from_status_id=open_status.id,
                    to_status_id=final_status.id,
                    changed_by=analista.id,
                    reason=MIGRATION_REASON,
                    created_at=cierre or inicio_atencion,
                )
            )
            comment_body = f"Canal: {row['canal']} · Empresa: {row['empresa']}\n\nSolución: {str(row['solucion']).strip()}"
            db.add(
                TicketComment(
                    ticket_id=ticket.id,
                    author_id=analista.id,
                    body=comment_body,
                    is_internal=True,
                    created_at=cierre or inicio_atencion,
                    updated_at=cierre or inicio_atencion,
                )
            )
            db.add(
                AuditLog(
                    user_id=None,
                    action="CREATE",
                    entity="Ticket",
                    entity_id=str(ticket.id),
                    new_values={"ticketNumber": ticket.ticket_number, "reason": "import_registro_tickets.py"},
                )
            )

        await db.commit()
        print(f"\nImportación completa: {len(rows_sorted)} tickets creados.")


if __name__ == "__main__":
    asyncio.run(main())
